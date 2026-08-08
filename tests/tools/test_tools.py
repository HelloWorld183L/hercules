from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook

from hercules.tools import (
    create_moving_avg_graph,
    create_volume_graph,
    describe_excel_workbook,
    query_sheet,
)


def _load_search_module_with_vector_client(monkeypatch, fake_vector_client):
    import hercules.tools.search_knowledgebase as search_module

    monkeypatch.setattr(search_module, "get_vector_client", lambda: fake_vector_client)
    return search_module


@pytest.fixture(autouse=True)
def use_agg_backend(monkeypatch):
    # Ensure matplotlib does not require a display in CI/test environments
    monkeypatch.setenv("MPLBACKEND", "Agg")
    yield


def test_create_moving_avg_graph_success():
    payload = {
        "toolUseId": "test-tool-use",
        "input": {
            "dates": ["2024-01-01", "2024-01-02", "2024-01-03"],
            "metrics": ["1", "2", "3"],
        },
    }

    result = create_moving_avg_graph.create_moving_avg_graph(payload)

    assert result["status"] == "success"
    assert result["toolUseId"] == "test-tool-use"
    assert isinstance(result["content"], list)
    assert result["content"][0]["image"]["format"] == "png"
    assert isinstance(result["content"][0]["image"]["source"]["bytes"], bytes)
    assert result["content"][0]["image"]["source"]["bytes"]


def test_create_moving_avg_graph_empty_dates_returns_error():
    payload = {
        "toolUseId": "test-tool-use",
        "input": {"dates": [], "metrics": ["1", "2"]},
    }

    result = create_moving_avg_graph.create_moving_avg_graph(payload)

    assert result["status"] == "error"
    assert "No dates provided" in result["content"][0]["text"]


def test_create_volume_graph_success():
    payload = {
        "toolUseId": "test-tool-use",
        "input": {
            "muscle_groups": ["legs", "back"],
            "volume_per_muscle_group": ["10", "20"],
            "workout_program_name": "Test Program",
        },
    }

    result = create_volume_graph.create_volume_graph(payload)

    assert result["status"] == "success"
    assert result["toolUseId"] == "test-tool-use"
    assert result["content"][0]["image"]["format"] == "png"
    assert isinstance(result["content"][0]["image"]["source"]["bytes"], bytes)


def test_create_volume_graph_invalid_volume_values_returns_error():
    payload = {
        "toolUseId": "test-tool-use",
        "input": {
            "muscle_groups": ["legs", "back"],
            "volume_per_muscle_group": ["10", "abc"],
        },
    }

    result = create_volume_graph.create_volume_graph(payload)

    assert result["status"] == "error"
    assert "Invalid volume values" in result["content"][0]["text"]


def test_describe_excel_workbook_returns_sheet_details(tmp_path: Path):
    workbook_path = tmp_path / "workbook.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Routine"
    sheet["A1"] = "Exercise"
    sheet["A2"] = "Squat"
    sheet["B2"] = "https://example.com"
    workbook.save(workbook_path)

    payload = {
        "toolUseId": "test-tool-use",
        "input": {"workbook_file_path": str(workbook_path)},
    }

    result = describe_excel_workbook.describe_excel_workbook(payload)

    assert result["status"] == "success"
    assert result["toolUseId"] == "test-tool-use"
    assert "sheets" in result["content"][1]["json"]
    assert "Routine" in result["content"][1]["json"]["sheets"]
    assert result["content"][1]["json"]["urls"] == ["https://example.com"]


def test_describe_excel_workbook_missing_path_returns_error():
    payload = {"toolUseId": "test-tool-use", "input": {"workbook_file_path": ""}}

    result = describe_excel_workbook.describe_excel_workbook(payload)

    assert result["status"] == "error"
    assert "No workbook data has been provided" in result["content"][0]["text"]


def test_query_sheet_returns_serialized_rows_and_urls(tmp_path: Path):
    workbook_path = tmp_path / "workbook.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Program"
    sheet["A1"] = "Visit"
    sheet["A2"] = "https://example.org"
    sheet["B1"] = "Sets"
    sheet["B2"] = 5
    workbook.save(workbook_path)

    payload = {
        "toolUseId": "test-tool-use",
        "input": {
            "workbook_file_path": str(workbook_path),
            "sheet_name": "Program",
        },
    }

    result = query_sheet.query_sheet(payload)

    assert result["status"] == "success"
    assert result["toolUseId"] == "test-tool-use"
    assert result["content"][1]["json"]["urls"] == ["https://example.org"]
    assert any(isinstance(row, list) for row in result["content"][1]["json"]["rows"])


def test_query_sheet_missing_sheet_returns_error(tmp_path: Path):
    workbook_path = tmp_path / "workbook.xlsx"
    workbook = Workbook()
    workbook.save(workbook_path)

    payload = {
        "toolUseId": "test-tool-use",
        "input": {
            "workbook_file_path": str(workbook_path),
            "sheet_name": "MissingSheet",
        },
    }

    result = query_sheet.query_sheet(payload)

    assert result["status"] == "error"
    assert "not found in workbook" in result["content"][0]["text"]


def test_search_knowledgebase_no_query_returns_error():
    import hercules.tools.search_knowledgebase as search_module

    payload = {"toolUseId": "test-tool-use", "input": {"query": ""}}

    result = search_module.search_knowledgebase(payload)

    assert result["status"] == "error"
    assert "No query provided" in result["content"][0]["text"]


def test_search_knowledgebase_returns_no_results(monkeypatch):
    class FakeResult:
        def __init__(self):
            self.points = []

    class FakeClient:
        def search(self, query, limit):
            return FakeResult()

    search_module = _load_search_module_with_vector_client(monkeypatch, FakeClient())

    payload = {"toolUseId": "test-tool-use", "input": {"query": "something"}}
    result = search_module.search_knowledgebase(payload)

    assert result["status"] == "success"
    assert (
        result["content"][0]["text"]
        == "No relevant documents found in the knowledge base."
    )


def test_search_knowledgebase_formats_results_from_payload_objects(monkeypatch):
    class FakePoint:
        def __init__(self, payload, score):
            self.payload = payload
            self.score = score

    class FakeClient:
        def search(self, query, limit):
            return SimpleNamespace(
                points=[FakePoint({"text": "Doc text", "source": "source.md"}, 0.85)]
            )

    search_module = _load_search_module_with_vector_client(monkeypatch, FakeClient())

    payload = {"toolUseId": "test-tool-use", "input": {"query": "something"}}
    result = search_module.search_knowledgebase(payload)

    assert result["status"] == "success"
    assert result["content"][0]["text"] == "Doc text"
    assert result["content"][0]["source"] == "source.md"
    assert result["content"][0]["score"] == 0.85


def test_search_knowledgebase_formats_results_from_tuple_payloads(monkeypatch):
    class FakeClient:
        def search(self, query, limit):
            return SimpleNamespace(
                points=[({"text": "Tuple text", "source": "tuple.md"}, 0.75)]
            )

    search_module = _load_search_module_with_vector_client(monkeypatch, FakeClient())

    payload = {"toolUseId": "test-tool-use", "input": {"query": "something"}}
    result = search_module.search_knowledgebase(payload)

    assert result["status"] == "success"
    assert result["content"][0]["text"] == "Tuple text"
    assert result["content"][0]["source"] == "tuple.md"
    assert result["content"][0]["score"] == 0.75
