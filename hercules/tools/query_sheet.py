import base64
import datetime as _dt
import logging
import re

from openpyxl import load_workbook
from strands.types.tools import ToolResult, ToolUse

from hercules.tools.helpers import extract_urls_from_rows

logger = logging.getLogger("hercules")

# Regex to extract http(s) URLs from text
URL_RE = re.compile(r"https?://[^\s,;\)\]\}'\"]+")
MAX_URLS_PER_SHEET = 50

TOOL_SPEC = {
    "name": "query_sheet",
    "description": "Get the relevant rows and columns from a worksheet to understand a workout program.",
    "inputSchema": {
        "json": {
            "type": "object",
            "properties": {
                "workbook_file_path": {
                    "type": "string",
                    "description": "File path for the temporary Excel workbook. It must include the file extension.",
                },
                "sheet_name": {
                    "type": "string",
                    "description": "Sheet name of interest within the Excel workbook.",
                },
            },
            "required": ["workbook_file_path", "sheet_name"],
        }
    },
}


def query_sheet(tool: ToolUse, **kwargs) -> ToolResult:
    """
    Get the relevant rows and columns from a worksheet to understand a workout program.
    Returns the relevant rows and columns from a worksheet to understand a workout program.
    """
    tool_use_id = tool["toolUseId"]
    tool_input = tool["input"]
    workbook_file_path = tool_input["workbook_file_path"]
    workbook_sheet_name = tool_input["sheet_name"]

    MAX_ROWS = 500
    MAX_COLS = 50

    if not workbook_file_path:
        return {
            "toolUseId": tool_use_id,
            "status": "error",
            "content": [{"text": "No workbook data has been provided."}],
        }

    workbook = load_workbook(filename=workbook_file_path, read_only=True)
    if workbook_sheet_name not in workbook.sheetnames:
        return {
            "toolUseId": tool_use_id,
            "status": "error",
            "content": [
                {"text": f"Sheet name {workbook_sheet_name} not found in workbook."}
            ],
        }

    # We set this in order to not overwhelm the context window with too much data
    # TODO: Look into context offloading via retrieval to be able to retrieve more data without overwhelming the context window
    sheet = workbook[workbook_sheet_name]
    sheet_rows = [
        row
        for row in sheet.iter_rows(max_row=MAX_ROWS, max_col=MAX_COLS, values_only=True)
    ]

    # Serialize cell values so the result is JSON-safe (datetimes -> ISO, bytes -> base64)
    # Convert to a sparse representation: keep only non-null cells as {col, value} entries
    serialized_rows = [
        [
            {"col": i + 1, "value": _serialize_cell(c)}
            for i, c in enumerate(row)
            if c is not None
        ]
        for row in sheet_rows
    ]
    logger.info("`query_sheet` tool: Retrieved and serialized rows.")
    # Extract URLs from the retrieved rows so the agent can decide to fetch them
    urls = extract_urls_from_rows(sheet_rows, max_urls_per_sheet=MAX_URLS_PER_SHEET)

    return {
        "toolUseId": tool_use_id,
        "status": "success",
        "content": [
            {
                "text": f"Workbook has been analysed successfully. The first {len(sheet_rows)} rows and up to {MAX_COLS} columns have been retrieved. Only non-empty cells are returned."
            },
            {"json": {"rows": serialized_rows, "urls": urls}},
        ],
    }


def _serialize_cell(cell):
    """
    Serialize Excel cell values to be JSON-safe.

    :param cell: The cell value to serialize.
    :returns: A JSON-safe representation of the cell value.

    """
    if isinstance(cell, (_dt.datetime, _dt.date, _dt.time)):
        return cell.isoformat()
    if isinstance(cell, (bytes, bytearray)):
        return base64.b64encode(cell).decode("ascii")
    return cell
