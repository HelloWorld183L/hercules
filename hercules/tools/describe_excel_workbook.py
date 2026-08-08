import logging
import re

from openpyxl import load_workbook
from strands.types.tools import ToolResult, ToolUse

from hercules.tools.helpers import extract_urls_from_rows

logger = logging.getLogger("hercules")

# Regex to extract http(s) URLs from text
URL_RE = re.compile(r"https?://[^\s,;\)\]\}'\"]+")
MAX_URLS_PER_SHEET = 50

TOOL_SPEC = {
    "name": "describe_excel_workbook",
    "description": "Describe the structure of an Excel workbook. These are usually workout programs where workout structure, progression methods, etc. are described. This is used to understand the layout and content of the workout program without reading the entire file contents. Returns workbook structure, sheet names, columns, row counts.",
    "inputSchema": {
        "json": {
            "type": "object",
            "properties": {
                "workbook_file_path": {
                    "type": "string",
                    "description": "File path for the temporary Excel workbook. It must include the file extension.",
                }
            },
            "required": ["workbook_file_path"],
        }
    },
}


def describe_excel_workbook(tool: ToolUse, **kwargs) -> ToolResult:
    """
    Describe the structure of an Excel workbook (`.xlsx` file format).
    These are usually workout programs where workout structure, progression methods, etc. are described.
    This is used to understand the layout and content of the workout program without reading the entire file contents.
    Returns workbook structure, sheet names, columns, row counts.
    """
    tool_use_id = tool["toolUseId"]
    tool_input = tool["input"]
    workbook_file_path = tool_input["workbook_file_path"]

    logger.info(f"Received request to describe workbook at {workbook_file_path}")

    if not workbook_file_path:
        return {
            "toolUseId": tool_use_id,
            "status": "error",
            "content": [{"text": "No workbook data has been provided."}],
        }

    SAMPLE_MAX_ROWS = 10
    SAMPLE_MAX_COLS = 20

    workbook = load_workbook(filename=workbook_file_path, read_only=True)
    logger.info(f"Workbook loaded successfully from {workbook_file_path}")

    workbook_sheet_names = workbook.sheetnames
    sheet_details = {}
    aggregated_urls = []
    aggregated_seen = set()
    for sheet_name in workbook_sheet_names:
        sheet = workbook[sheet_name]
        sample_rows = [
            row
            for row in sheet.iter_rows(
                max_row=SAMPLE_MAX_ROWS, max_col=SAMPLE_MAX_COLS, values_only=True
            )
        ]

        urls = extract_urls_from_rows(
            sample_rows, max_urls_per_sheet=MAX_URLS_PER_SHEET
        )
        for u in urls:
            if u not in aggregated_seen:
                aggregated_seen.add(u)
                aggregated_urls.append(u)

        sheet_details[sheet_name] = {
            "row_count": sheet.max_row,
            "column_count": sheet.max_column,
            # Sample the sheet to get an idea as to what is in it
            "sample_rows": sample_rows,
            "urls": urls,
        }

    if len(sheet_details) == 0:
        return {
            "toolUseId": tool_use_id,
            "status": "error",
            "content": [{"text": "No sheets found in the workbook."}],
        }

    logger.info(f"Workbook description generated successfully for {workbook_file_path}")

    return {
        "toolUseId": tool_use_id,
        "status": "success",
        "content": [
            {
                "text": f"{workbook_file_path} workbook structure has been analysed successfully. Contents have been sampled with a maximum of {SAMPLE_MAX_ROWS} rows and {SAMPLE_MAX_COLS} columns."
            },
            {"json": {"sheets": sheet_details, "urls": aggregated_urls}},
        ],
    }
